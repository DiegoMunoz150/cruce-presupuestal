import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference  # Importación para el gráfico
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.drawing.line import LineProperties

def generar_cruce(archivo_ingresos, archivo_egresos, archivo_salida):
    
    wb_ingresos = load_workbook(archivo_ingresos, data_only=False)
    wb_egresos = load_workbook(archivo_egresos, data_only=False)
    ws_ingresos = wb_ingresos.active
    ws_egresos = wb_egresos.active

    wb_nuevo = Workbook()
    
    # 1. HOJA 1: INGRESOS
    ws1 = wb_nuevo.active
    ws1.title = "INGRESOS"
    for fila in ws_ingresos.iter_rows():
        for celda in fila:
            ws1[celda.coordinate] = celda.value

    # 2. HOJA 2: EGRESOS
    ws2 = wb_nuevo.create_sheet("EGRESOS")
    for fila in ws_egresos.iter_rows():
        for celda in fila:
            ws2[celda.coordinate] = celda.value

    # 3. HOJA 3: CRUCE INFORMACION
    ws3 = wb_nuevo.create_sheet("CRUCE INFORMACION")
    encabezados = [
        "FUENTE DE RECURSO", "NOMBRE DE FUENTE DE RECURSO", 
        "PRESUPUESTO DEFINITIVO DE INGRESOS", "PRESUPUESTO DEFINITIVO DE GASTOS", 
        "DIFERENCIA", "Recaudo", "Registros", "Saldos no Comprometidos", 
        "Constitución", "Reservas", "Pago", "Cuentas x Pagar", "Recurso en Bancos"
    ]

    gris_encabezado = "7F7F7F"
    azul_oscuro = "385D8A"
    morado = "7030A0"
    azul = "0070C0"
    naranja = "ED7D31"
    verde_encabezado = "92D050"

    colores = [gris_encabezado, gris_encabezado, azul_oscuro, azul_oscuro, azul_oscuro, morado, morado, morado, azul, azul, naranja, naranja, verde_encabezado]

    for col, (titulo, color) in enumerate(zip(encabezados, colores), start=1):
        celda = ws3.cell(row=1, column=col)
        celda.value = titulo
        celda.fill = PatternFill("solid", fgColor=color)
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    ws3.row_dimensions[1].height = 30

    encabezados_egresos = {}
    for col in range(1, ws2.max_column + 1):
        valor = ws2.cell(row=2, column=col).value
        if valor is not None:
            encabezados_egresos[str(valor).strip().upper()] = col

    col_fuente = encabezados_egresos["FUENTE DE RECURSO"]
    col_nombre = encabezados_egresos["NOMBRE DE FUENTE DE RECURSO"]

    fuentes_vistas = set()
    fila_destino = 2

    for fila in range(3, ws2.max_row + 1):
        fuente = ws2.cell(row=fila, column=col_fuente).value
        nombre = ws2.cell(row=fila, column=col_nombre).value
        if fuente is None or str(fuente).strip() == "":
            continue
        try:
            if float(fuente) == 0:
                continue
        except:
            pass

        clave = str(fuente).strip()
        if clave in fuentes_vistas:
            continue

        fuentes_vistas.add(clave)
        ws3.cell(row=fila_destino, column=1, value=fuente)
        ws3.cell(row=fila_destino, column=2, value=nombre)
        fila_destino += 1

    ultima_fila = ws3.max_row
    for fila in range(2, ultima_fila + 1):
        ws3[f"C{fila}"] = f'=SUMIF(INGRESOS!$R:$R,A{fila},INGRESOS!$L:$L)'
        ws3[f"D{fila}"] = f'=SUMIF(EGRESOS!$P:$P,A{fila},EGRESOS!$Y:$Y)'
        ws3[f"E{fila}"] = f'=C{fila}-D{fila}'
        ws3[f"F{fila}"] = f'=SUMIF(INGRESOS!$R:$R,A{fila},INGRESOS!$N:$N)'
        ws3[f"G{fila}"] = f'=SUMIF(EGRESOS!$P:$P,A{fila},EGRESOS!$AA:$AA)'
        ws3[f"H{fila}"] = f'=F{fila}-G{fila}'
        ws3[f"I{fila}"] = f'=SUMIF(EGRESOS!$P:$P,A{fila},EGRESOS!$AB:$AB)'
        ws3[f"J{fila}"] = f'=G{fila}-I{fila}'
        ws3[f"K{fila}"] = f'=SUMIF(EGRESOS!$P:$P,A{fila},EGRESOS!$AC:$AC)'
        ws3[f"L{fila}"] = f'=I{fila}-K{fila}'
        ws3[f"M{fila}"] = f'=H{fila}+J{fila}+L{fila}'

    # ==========================================================================
    # 4. HOJA 4: REPORTE FILTRADO + GRÁFICO PERSONALIZADO (image_5b0a1e.png)
    # ==========================================================================
    ws4 = wb_nuevo.create_sheet("REPORTE FILTRADO")
    
    encabezados_filtrados = ["FUENTE DE RECURSO", "NOMBRE DE FUENTE DE RECURSO", "PRESUPUESTO DEFINITIVO", "RECAUDOS"]
    for col, titulo in enumerate(encabezados_filtrados, start=1):
        celda = ws4.cell(row=1, column=col, value=titulo)
        celda.fill = PatternFill("solid", fgColor="1F497D") 
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")
    
    codigos_objetivo = [
        "1.1.01.01.200", "1.1.01.02.200", "1.1.01.02.204", 
        "1.1.01.02.218", "1.1.01.02.300.01", "1.1.01.02.300.55", "1.1.01.02.300.61"
    ]
    
    thin = Side(style="thin", color="000000")
    fila_filt = 2
    
    encabezados_ing = {}
    for col in range(1, ws1.max_column + 1):
        v = ws1.cell(row=2, column=col).value 
        if v is not None:
            encabezados_ing[str(v).strip().upper()] = col
            
    idx_rubro = encabezados_ing.get("RUBRO", 1)
    idx_nombre = encabezados_ing.get("NOMBRE DE FUENTE DE RECURSO", 2) 

    for fila in range(3, ws1.max_row + 1):
        rubro_val = str(ws1.cell(row=fila, column=idx_rubro).value).strip()
        
        if rubro_val in codigos_objetivo:
            ws4.cell(row=fila_filt, column=1, value=ws1.cell(row=fila, column=idx_rubro).value)
            ws4.cell(row=fila_filt, column=2, value=ws1.cell(row=fila, column=idx_nombre).value)
            ws4.cell(row=fila_filt, column=3, value=f'=INGRESOS!L{fila}')
            ws4.cell(row=fila_filt, column=4, value=f'=INGRESOS!N{fila}')
            
            ws4[f"C{fila_filt}"].number_format = '$#,##0.00'
            ws4[f"D{fila_filt}"].number_format = '$#,##0.00'
            fila_filt += 1

    for r in range(1, ws4.max_row + 1):
        for c in range(1, 5):
            ws4.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


# --------------------------------------------------------------------------
    # CONSTRUCCIÓN DEL GRÁFICO DE BARRAS HORIZONTALES (CORREGIDO Y LIMPIO)
    # --------------------------------------------------------------------------
    if ws4.max_row > 1:
        chart = BarChart()
        chart.type = "bar"              
        chart.style = 10                
        chart.title = "Esfuerzo tributario municipal"
        
        # Referencias de datos y categorías
        data = Reference(ws4, min_col=3, min_row=1, max_col=4, max_row=ws4.max_row)
        cats = Reference(ws4, min_col=2, min_row=2, max_row=ws4.max_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # --- SOLUCIÓN AL AMONTONAMIENTO DE TEXTO ---
        chart.dataLabels = openpyxl.chart.label.DataLabelList()
        chart.dataLabels.showVal = True          # SÍ muestra el número
        chart.dataLabels.showSerName = False      # NO muestra "PRESUPUESTO" o "RECAUDOS" en cada barra
        chart.dataLabels.showCatName = False      # NO repite el nombre del impuesto en la barra
        chart.dataLabels.showPercent = False      # NO muestra porcentajes
        
        # Formato numérico para las etiquetas dentro del gráfico (formato contable compacto)
        chart.dataLabels.numFmt = "$#,##0" 

        # --- SOLUCIÓN A LA ORIENTACIÓN E INVERSIÓN DEL EJE LEFT ---
        chart.y_axis.tickLblPos = "low"            # Nombres fijos estrictamente a la izquierda
        chart.y_axis.scaling.orientation = "maxMin" # Invierte el orden para que lea de arriba a abajo como tu tabla
        
        # Dimensiones amplias para que respiren las etiquetas de dinero
        chart.height = 14
        chart.width = 25  
        
        # Colores idénticos a tu paleta institucional
        if len(chart.series) >= 2:
            # Serie 1: Presupuesto Definitivo (Gris)
            chart.series[0].graphicalProperties.solidFill = "7F7F7F"
            # Serie 2: Recaudos (Verde)
            chart.series[1].graphicalProperties.solidFill = "00B050"

        # Inyectar el gráfico limpio al lado de la tabla
        ws4.add_chart(chart, "F2")

    # Autoajuste de columnas de la hoja 4
    for col in ws4.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws4.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ==========================================================================
    # 5. HOJA 5: INFORME RAPIDO (Migrado)
    # ==========================================================================
    ws5 = wb_nuevo.create_sheet("INFORME RAPIDO")

    ws5["A1"] = "CONCEPTO"
    ws5["B1"] = "VALOR"
    ws5["A2"] = "Recaudo hoja INGRESOS"
    ws5["B2"] = "=INGRESOS!N2"
    ws5["A3"] = "Suma recaudo hoja CRUCE INFORMACION"
    ws5["B3"] = "=SUM('CRUCE INFORMACION'!F:F)"
    ws5["A4"] = "Diferencia"
    ws5["B4"] = "=B2-B3"
    ws5["A5"] = "Resultado"
    ws5["B5"] = (
        '=IF(ABS(B4)<0.01,'
        '"Los datos coinciden",'
        '"Los datos no coinciden, faltan productos")'
    )

    ws5["A8"] = "VALIDACION PRESUPUESTO DEFINITIVO"
    ws5["A9"] = "Presupuesto Definitivo INGRESOS"
    ws5["B8"] = "VALOR"
    ws5["B9"] = "=INGRESOS!L2"
    ws5["A10"] = "Suma Presupuesto Definitivo CRUCE"
    ws5["B10"] = "=SUM('CRUCE INFORMACION'!C:C)"
    ws5["A11"] = "Diferencia"
    ws5["B11"] = "=B9-B10"
    ws5["A12"] = "Resultado"
    ws5["B12"] = (
        '=IF(ABS(B11)<0.01,'
        '"Los datos coinciden",'
        '"Los datos no coinciden, falta un producto")'
    )

    for celda in ["B2", "B3", "B4", "B9", "B10", "B11"]:
        ws5[celda].number_format = '$#,##0.00'

    for fila in range(1, 6):
        for col in range(1, 3):
            ws5.cell(fila, col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for fila in range(8, 13):
        for col in range(1, 3):
            ws5.cell(fila, col).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws5.column_dimensions["A"].width = 40
    ws5.column_dimensions["B"].width = 25

    # ==========================================================================
    # REGLAS GENERALES Y FORMATOS DE BORDES RESTANTES
    # ==========================================================================
    for fila in range(1, ws3.max_row + 1):
        for col in range(1, ws3.max_column + 1):
            celda = ws3.cell(row=fila, column=col)
            celda.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(3, 14):  
        letra_col = get_column_letter(col)
        for fila in range(2, ws3.max_row + 1):
            ws3[f"{letra_col}{fila}"].number_format = '$#,##0.00'
            
    for col in ws3.columns:
        max_length = 0
        columna = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws3.column_dimensions[columna].width = min(max_length + 3, 40)

    ws3.conditional_formatting.add(
        f"C2:M{ws3.max_row}",
        CellIsRule(
            operator='lessThan',
            formula=['0'],
            fill=PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        )
    )


# ==========================================================================
    # NUEVA HOJA "test": DASHBOARD VISUAL (GRÁFICA ESTILO POWER BI)
    # ==========================================================================
    ws_test = wb_nuevo.create_sheet("test")
    
    # TRUCO DE DISEÑO 1: Apagar las celdas de Excel para que parezca un lienzo en blanco
    ws_test.sheet_view.showGridLines = False

    if ws4.max_row > 1:
        chart_test = BarChart()
        chart_test.type = "bar"              
        chart_test.style = 2  # Estilo más limpio y moderno
        chart_test.title = "📊 Resumen de Esfuerzo Tributario Municipal"
        
        # El gráfico vive en 'test', pero chupa los datos de la hoja 'REPORTE FILTRADO' (ws4)
        data_test = Reference(ws4, min_col=3, min_row=1, max_col=4, max_row=ws4.max_row)
        cats_test = Reference(ws4, min_col=2, min_row=2, max_row=ws4.max_row)
        
        chart_test.add_data(data_test, titles_from_data=True)
        chart_test.set_categories(cats_test)
        
        # --- DISEÑO LIMPIO (ESTILO DASHBOARD) ---
        
        # Poner la leyenda arriba para que no estorbe a los lados
        chart_test.legend.position = "t"

        # TRUCO DE DISEÑO 2: Quitar la línea negra que envuelve al gráfico
        chart_test.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
        

        # Etiquetas de datos elegantes (los números al lado de las barras)
        chart_test.dataLabels = openpyxl.chart.label.DataLabelList()
        chart_test.dataLabels.showVal = True          
        chart_test.dataLabels.numFmt = "$#,##0" 
        
        # TRUCO DE DISEÑO 3: Limpiar el eje horizontal para no saturar la vista
        chart_test.x_axis.majorGridlines = None  # Quita las líneas verticales de fondo
        chart_test.x_axis.delete = False         # Mantiene los números de abajo visibles
        
        # Mantener los nombres alineados a la izquierda y en orden lógico
        chart_test.y_axis.tickLblPos = "low"            
        chart_test.y_axis.scaling.orientation = "maxMin" 
        
        # Tamaño panorámico (más grande para la hoja completa)
        chart_test.height = 16
        chart_test.width = 30  
        
        # Colores ajustados para mayor contraste
        if len(chart_test.series) >= 2:
            # Presupuesto Definitivo: Gris más suave, menos agresivo
            chart_test.series[0].graphicalProperties.solidFill = "D9D9D9"
            # Recaudos: Verde institucional fuerte
            chart_test.series[1].graphicalProperties.solidFill = "00B050"

        # Insertar el gráfico flotando en el lienzo en blanco
        ws_test.add_chart(chart_test, "B2")
        
    # ==========================================================================
    # HOJA NUEVA: GRAFICOS - "Esfuerzo tributario municipal" (versión mejorada)
    # ==========================================================================
        ws_graf = wb_nuevo.create_sheet("GRAFICOS")

        chart2 = BarChart()
        chart2.type = "bar"                 # barras horizontales
        chart2.grouping = "clustered"       # agrupadas, NO superpuestas
        chart2.overlap = -20                # separación entre series (negativo = espacio)
        chart2.gapWidth = 100                # espacio entre categorías
        chart2.style = 10
        chart2.title = "Esfuerzo tributario municipal"

        # Datos: Presupuesto Definitivo (col 3) y Recaudos (col 4)
        data2 = Reference(ws4, min_col=3, min_row=1, max_col=4, max_row=ws4.max_row)
        cats2 = Reference(ws4, min_col=2, min_row=2, max_row=ws4.max_row)

        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)

        # --- Etiquetas de datos: solo el valor, formato compacto ---
        chart2.dataLabels = openpyxl.chart.label.DataLabelList()
        chart2.dataLabels.showVal = True
        chart2.dataLabels.showSerName = False
        chart2.dataLabels.showCatName = False
        chart2.dataLabels.showPercent = False
        chart2.dataLabels.showLegendKey = False
        chart2.dataLabels.numFmt = "$#,##0"
        chart2.dataLabels.position = "outEnd"   # etiqueta fuera de la barra, evita solaparse

        # --- Eje Y: nombres a la izquierda, orden de arriba a abajo igual a la tabla ---
        chart2.y_axis.tickLblPos = "low"
        chart2.y_axis.scaling.orientation = "maxMin"
        chart2.y_axis.delete = False

        # --- Eje X (valores) ---
        chart2.x_axis.numFmt = "$#,##0"
        chart2.x_axis.delete = False

        # --- Leyenda: arriba, fuera del área de barras ---
        chart2.legend.position = "t"
        chart2.legend.overlay = False

        # --- Colores institucionales ---
        if len(chart2.series) >= 2:
            chart2.series[0].graphicalProperties.solidFill = "7F7F7F"  # Presupuesto Definitivo - gris
            chart2.series[1].graphicalProperties.solidFill = "00B050"  # Recaudos - verde

        # --- Dimensiones amplias para que respiren etiquetas y nombres ---
        chart2.height = 16
        chart2.width = 28

        ws_graf.add_chart(chart2, "B2")


    # ==========================================================================
    # HOJA NUEVA: GRAFICOS - "Informe de Recaudo - Principales Ingresos Tributarios"
    # ==========================================================================
    ws_graf2 = wb_nuevo.create_sheet("GRAF PRINCIPALES IMPUESTOS")

    ws_graf2["A1"] = "NOMBRE"
    ws_graf2["B1"] = "PRESUPUESTO DEFINITIVO"
    ws_graf2["C1"] = "RECAUDOS"

    impuestos_principales = [
        "IMPUESTO PREDIAL UNIFICADO",
        "SOBRETASA A LA GASOLINA",
        "IMPUESTO DE INDUSTRIA Y COMERCIO",
    ]

    # Encabezados de INGRESOS están en la FILA 1 (no la 2)
    encabezados_ing = {}
    for col in range(1, ws1.max_column + 1):
        v = ws1.cell(row=1, column=col).value
        if v is not None:
            encabezados_ing[str(v).strip().upper()] = col

    idx_nombre_ing = encabezados_ing.get("NOMBRE")              # col B
    idx_presup_ing = encabezados_ing.get("PRESUPUESTO  DEFINITIVO")  # OJO: doble espacio, col L
    idx_recaudo_ing = encabezados_ing.get("RECAUDOS")           # col N

    fila_g2 = 2
    for nombre_obj in impuestos_principales:
        for fila in range(2, ws1.max_row + 1):
            nombre_val = ws1.cell(row=fila, column=idx_nombre_ing).value
            if nombre_val is not None and str(nombre_val).strip().upper() == nombre_obj:
                col_p = get_column_letter(idx_presup_ing)
                col_r = get_column_letter(idx_recaudo_ing)
                ws_graf2.cell(row=fila_g2, column=1, value=nombre_val)
                ws_graf2.cell(row=fila_g2, column=2, value=f"=INGRESOS!{col_p}{fila}")
                ws_graf2.cell(row=fila_g2, column=3, value=f"=INGRESOS!{col_r}{fila}")
                ws_graf2[f"B{fila_g2}"].number_format = '$#,##0.00'
                ws_graf2[f"C{fila_g2}"].number_format = '$#,##0.00'
                fila_g2 += 1
                break

    # --- Gráfico de columnas verticales ---
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "clustered"
    chart3.overlap = -10
    chart3.gapWidth = 150
    chart3.style = 10
    chart3.title = "Informe de Recaudo - Principales Ingresos Tributarios"

    data3 = Reference(ws_graf2, min_col=2, min_row=1, max_col=3, max_row=fila_g2 - 1)
    cats3 = Reference(ws_graf2, min_col=1, min_row=2, max_row=fila_g2 - 1)

    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)

    chart3.dataLabels = openpyxl.chart.label.DataLabelList()
    chart3.dataLabels.showVal = True
    chart3.dataLabels.showSerName = False
    chart3.dataLabels.showCatName = False
    chart3.dataLabels.showPercent = False
    chart3.dataLabels.showLegendKey = False
    chart3.dataLabels.numFmt = '$#,##0.00'
    chart3.dataLabels.position = "outEnd"

    chart3.x_axis.delete = False
    chart3.y_axis.delete = True

    chart3.legend.position = "b"
    chart3.legend.overlay = False

    if len(chart3.series) >= 2:
        chart3.series[0].graphicalProperties.solidFill = "1F77B4"
        chart3.series[1].graphicalProperties.solidFill = "BFBFBF"

    chart3.height = 10
    chart3.width = 22

    ws_graf2.add_chart(chart3, "E2")


    wb_nuevo.save(archivo_salida)
    return archivo_salida