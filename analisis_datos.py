
import pandas as pd
import openpyxl

def cargar_ingresos(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    data = []
    for fila in range(2, ws.max_row + 1):
        data.append([ws.cell(row=fila, column=c).value for c in range(1, ws.max_column+1)])
    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df

def top_impuestos_tributarios(df, n=3):
    """
    Devuelve los N rubros tributarios más importantes por Presupuesto Definitivo.
    Filtra solo rubros 'hoja' (sin hijos directos) dentro de 1.1.01 (Ingresos tributarios).
    """
    df = df.copy()
    df["RUBRO"] = df["RUBRO"].astype(str)
    
    # Solo tributarios
    tributarios = df[df["RUBRO"].str.startswith("1.1.01.0")]
    
    # Filtrar rubros "hoja": ningún otro rubro empieza con este + "."
    rubros = set(tributarios["RUBRO"])
    def es_hoja(rubro):
        return not any(r.startswith(rubro + ".") for r in rubros)
    
    hojas = tributarios[tributarios["RUBRO"].apply(es_hoja)]
    
    # Top N por presupuesto definitivo
    col_presup = "PRESUPUESTO  DEFINITIVO"  # doble espacio
    top = hojas.nlargest(n, col_presup)[["RUBRO", "NOMBRE", col_presup, "RECAUDOS"]]
    return top

print(top_impuestos_tributarios(cargar_ingresos(r"C:\Users\USER\Desktop\Proyecto_DiegoMuñoz\cruce-presupuestal\uploads\ingresos.xlsx")))

